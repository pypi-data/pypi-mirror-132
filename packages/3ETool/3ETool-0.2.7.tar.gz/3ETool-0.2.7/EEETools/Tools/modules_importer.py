from EEETools.MainModules.main_module import CalculationOptions
from EEETools.Tools.Other.fernet_handler import FernetHandler
from openpyxl import Workbook, load_workbook, styles, utils
from EEETools.MainModules.main_module import ArrayHandler
from datetime import date, datetime
import math, pandas, os


def calculate_excel(excel_path, calculation_option=None):

    array_handler = import_excel_input(excel_path)
    array_handler.calculate()

    if calculation_option is not None and type(calculation_option) == CalculationOptions:

        array_handler.options = calculation_option

    export_solution_to_excel(excel_path, array_handler)


def calculate_dat(dat_path):

    array_handler = import_dat(dat_path)
    array_handler.calculate()
    write_csv_solution(dat_path, array_handler)


def convert_excel_to_dat(excel_path: str):

    array_handler = import_excel_input(excel_path)

    if ".xlsm" in excel_path:

        dat_path = excel_path.replace(".xlsm", ".dat")

    elif ".xlsx" in excel_path:

        dat_path = excel_path.replace(".xlsx", ".dat")

    else:

        dat_path = excel_path.replace(".xls", ".dat")

    export_dat(dat_path, array_handler)


def import_excel_input(excel_path) -> ArrayHandler:

    array_handler = ArrayHandler()

    # import connections
    excel_connection_data = pandas.read_excel(excel_path, sheet_name="Stream")

    for line in excel_connection_data.values:

        line = line.tolist()
        if not math.isnan(line[0]):
            new_conn = array_handler.append_connection()

            new_conn.index = line[0]
            new_conn.name = str(line[1])
            new_conn.exergy_value = line[2]

    # import blocks
    excel_block_data = pandas.read_excel(excel_path, sheet_name="Componenti")

    for line in excel_block_data.values:

        line = line.tolist()

        if not (math.isnan(line[0]) or type(line[0]) is str):

            if line[0] > 0:

                if "Heat Exchanger" in str(line[2]) or "Scambiatore" in str(line[2]):

                    new_block = array_handler.append_block("Heat Exchanger")
                    excel_connection_list = list()
                    excel_connection_list.append(str(line[2]))
                    excel_connection_list.extend(line[5:])

                else:

                    new_block = array_handler.append_block(str(line[2]))
                    excel_connection_list = line[5:]

                new_block.index = line[0]
                new_block.name = str(line[1])
                new_block.comp_cost = line[3]

                new_block.initialize_connection_list(excel_connection_list)

            else:

                array_handler.append_excel_costs_and_useful_output(line[5:-1], line[0] == 0, line[3])

    return array_handler


def export_solution_to_excel(excel_path, array_handler: ArrayHandler):

    result_df = get_result_data_frames(array_handler)

    # generation of time stamps for excel sheet name
    today = date.today()
    now = datetime.now()
    today_str = today.strftime("%d %b")
    now_str = now.strftime("%H.%M")

    for key in result_df.keys():
        __write_excel_file(excel_path, sheet_name=(key + " - " + today_str + " - " + now_str),
                           data_frame=result_df[key])


def export_dat(dat_path, array_handler: ArrayHandler):

    fernet = FernetHandler()
    fernet.save_file(dat_path, array_handler.xml)


def import_dat(dat_path) -> ArrayHandler:
    array_handler = ArrayHandler()
    fernet = FernetHandler()
    root = fernet.read_file(dat_path)
    array_handler.xml = root

    return array_handler


def write_csv_solution(dat_path, array_handler):
    result_df = get_result_data_frames(array_handler)

    # generation of time stamps for excel sheet name
    today = date.today()
    now = datetime.now()
    today_str = today.strftime("%d %b")
    now_str = now.strftime("%H.%M")

    dir_path = os.path.dirname(dat_path)

    for key in result_df.keys():
        csv_path = key + " - " + today_str + " - " + now_str + ".csv"
        csv_path = os.path.join(dir_path, csv_path)

        pandas_df = pandas.DataFrame(data=result_df[key])
        pandas_df.to_csv(path_or_buf=csv_path, sep="\t")


def get_result_data_frames(array_handler: ArrayHandler) -> dict:
    # Stream Solution Data frame generation
    stream_data = {"Stream": list(),
                   "Name": list(),
                   "Exergy Value [kW]": list(),
                   "Specific Cost [€/kJ]": list(),
                   "Specific Cost [€/kWh]": list(),
                   "Total Cost [€/s]": list()}

    for conn in array_handler.connection_list:

        if not conn.is_internal_stream:
            stream_data["Stream"].append(conn.index)
            stream_data["Name"].append(conn.name)
            stream_data["Exergy Value [kW]"].append(conn.exergy_value)
            stream_data["Specific Cost [€/kJ]"].append(conn.rel_cost)
            stream_data["Specific Cost [€/kWh]"].append(conn.rel_cost * 3600)
            stream_data["Total Cost [€/s]"].append(conn.rel_cost * conn.exergy_value)

    # Components Data frame generation
    comp_data = {"Name": list(),
                 "Comp Cost [€/s]": list(),

                 "Exergy_fuel [kW]": list(),
                 "Exergy_product [kW]": list(),
                 "Exergy_destruction [kW]": list(),
                 "Exergy_loss [kW]": list(),
                 "Exergy_dl [kW]": list(),

                 "Fuel Cost [€/kWh]": list(),
                 "Fuel Cost [€/s]": list(),
                 "Product Cost [€/kWh]": list(),
                 "Product Cost [€/s]": list(),
                 "Destruction Cost [€/kWh]": list(),
                 "Destruction Cost [€/s]": list(),

                 "eta": list(),
                 "r": list(),
                 "f": list(),
                 "y": list()}

    for block in array_handler.block_list:

        if not block.is_support_block:
            comp_data["Name"].append(block.name)
            comp_data["Comp Cost [€/s]"].append(block.comp_cost)

            comp_data["Exergy_fuel [kW]"].append(block.exergy_analysis["fuel"])
            comp_data["Exergy_product [kW]"].append(block.exergy_analysis["product"])
            comp_data["Exergy_destruction [kW]"].append(block.exergy_analysis["distruction"])
            comp_data["Exergy_loss [kW]"].append(block.exergy_analysis["losses"])
            comp_data["Exergy_dl [kW]"].append(block.exergy_analysis["distruction"] + block.exergy_analysis["losses"])

            try:

                comp_data["Fuel Cost [€/kWh]"].append(block.coefficients["c_fuel"] * 3600)
                comp_data["Product Cost [€/kWh]"].append(block.output_cost * 3600)
                comp_data["Destruction Cost [€/kWh]"].append(block.coefficients["c_dest"] * 3600)

                comp_data["Fuel Cost [€/s]"].append(block.coefficients["c_fuel"] * block.exergy_analysis["fuel"])
                comp_data["Product Cost [€/s]"].append(block.output_cost * block.exergy_analysis["fuel"])
                comp_data["Destruction Cost [€/s]"].append(
                    block.coefficients["c_dest"] * (block.exergy_analysis["distruction"] + block.exergy_analysis["losses"]))

                comp_data["eta"].append(block.coefficients["eta"])
                comp_data["r"].append(block.coefficients["r"])
                comp_data["f"].append(block.coefficients["f"])
                comp_data["y"].append(block.coefficients["y"])

            except:

                comp_data["Fuel Cost [€/kWh]"].append(0)
                comp_data["Product Cost [€/kWh]"].append(0)
                comp_data["Destruction Cost [€/kWh]"].append(0)

                comp_data["Fuel Cost [€/s]"].append(0)
                comp_data["Product Cost [€/s]"].append(0)
                comp_data["Destruction Cost [€/s]"].append(0)

                comp_data["eta"].append(0)
                comp_data["r"].append(0)
                comp_data["f"].append(0)
                comp_data["y"].append(0)

    # Output Stream Data frame generation
    useful_data = {"Stream": list(),
                   "Name": list(),
                   "Exergy Value [kW]": list(),
                   "Specific Cost [€/kJ]": list(),
                   "Specific Cost [€/kWh]": list(),
                   "Total Cost [€/s]": list()}

    for conn in array_handler.useful_effect_connections:
        useful_data["Stream"].append(conn.index)
        useful_data["Name"].append(conn.name)
        useful_data["Exergy Value [kW]"].append(conn.exergy_value)
        useful_data["Specific Cost [€/kJ]"].append(conn.rel_cost)
        useful_data["Specific Cost [€/kWh]"].append(conn.rel_cost * 3600)
        useful_data["Total Cost [€/s]"].append(conn.rel_cost * conn.exergy_value)

    return {"Stream Out": stream_data,
            "Comp Out": comp_data,
            "Eff Out": useful_data}


def __convert_result_data_frames(data_frame: dict) -> dict:

    new_data_frame = dict()

    for key in data_frame.keys():

        sub_dict = __get_sub_dict(key)
        sub_dict.update({"values": list()})

        if not sub_dict["name"] in new_data_frame.keys():

            new_data_frame.update({sub_dict["name"]: sub_dict})

        else:

            (new_data_frame[sub_dict["name"]])["unit"].append(sub_dict["unit"][0])

        (new_data_frame[sub_dict["name"]])["values"].append([data_frame[key]])

    return new_data_frame


def __get_sub_dict(key):

    if "[" in key:

        parts = key.split("[")

        name = parts[0].replace("[", "")
        measure_unit = [parts[1].split("]")[0].replace("]", "")]

    else:

        name = key
        measure_unit = list()

    return {"name": name, "unit": measure_unit}


def __write_excel_file(excel_path, sheet_name, data_frame: dict):

    data_frame = __convert_result_data_frames(data_frame)

    if not os.path.isfile(excel_path):

        wb = Workbook()

    else:

        wb = load_workbook(excel_path)

    if not sheet_name in wb.sheetnames:
        wb.create_sheet(sheet_name)

    sheet = wb[sheet_name]

    col = 2
    for key in data_frame.keys():

        row = 2
        sub_data_frame = data_frame[key]
        n_sub_element = len(sub_data_frame["unit"])

        if key == "Name":

            column_dimension = 35

        elif key == "Stream":

            column_dimension = 8

        else:

            column_dimension = 20

        if n_sub_element == 0:

            sheet.merge_cells(start_row=row, start_column=col, end_row=row + 1, end_column=col)
            n_sub_element = 1

        else:

            if n_sub_element > 1:
                sheet.merge_cells(start_row=row, start_column=col, end_row=row, end_column=col + n_sub_element - 1)

            for n in range(n_sub_element):
                row = 3
                cell = sheet.cell(row, col + n, value="[" + sub_data_frame["unit"][n] + "]")
                cell.alignment = styles.Alignment(horizontal="center", vertical="center")
                cell.font = styles.Font(italic=True, size=10)

        cell = sheet.cell(2, col, value=key)
        cell.alignment = styles.Alignment(horizontal="center", vertical="center")
        cell.font = styles.Font(bold=True)

        for n in range(n_sub_element):

            row = 5
            data_list = (sub_data_frame["values"])[n][0]
            sheet.column_dimensions[utils.get_column_letter(col)].width = column_dimension

            for data in data_list:
                sheet.cell(row, col, value=data)
                row += 1

            col += 1

    wb.save(excel_path)
