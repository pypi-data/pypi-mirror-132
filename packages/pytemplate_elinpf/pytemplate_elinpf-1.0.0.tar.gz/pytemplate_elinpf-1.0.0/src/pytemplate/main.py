from __future__ import annotations
from openpyxl import load_workbook
from typing import Dict
import os
import sys


def netmask_to_bit_length(netmask: str) -> int:
    # 掩码到掩码长度
    """
    >>> netmask_to_bit_length('255.255.255.0')
    24
    >>>
    """
    # 分割字符串格式的子网掩码为四段列表
    # 计算二进制字符串中 '1' 的个数
    # 转换各段子网掩码为二进制, 计算十进制
    return sum([bin(int(i)).count('1') for i in netmask.split('.')])


def wildcard_mask_to_netmask(wildcard_mask: str) -> str:
    """
    反掩码到掩码
    >>> wildcard_mask_to_netmask('0.0.31.255')
    '255.255.224.0'
    >>>
    """
    return '.'.join([str(int(255) - int(i)) for i in wildcard_mask.split('.')])


def netmask_to_wildcard_mask(netmask: str) -> str:
    """
    掩码到反掩码
    >>> netmask_to_wildcard_mask('255.255.224.0')
    '0.0.31.255'
    """
    return wildcard_mask_to_netmask(netmask)


def addr2dec(addr: str) -> int:
    """
    将IP地址转换为整数
    >>> addr2dec('192.168.0.1')
    3232235777
    """
    items = [int(x) for x in addr.split(".")]
    return sum([items[i] << [24, 16, 8, 0][i] for i in range(4)])


def dec2addr(dec: int) -> str:
    """
    将整数恢复为IP地址
    >>> dec2addr(3232235777)
    '192.168.0.1'
    """
    return ".".join([str(dec >> x & 0xff) for x in [24, 16, 8, 0]])


def add_ip_address(ip: str, add_value: int) -> str:
    """
    简易的地址加法，无法判断网段
    >>> add_ip_address('192.168.0.1', 1)
    '192.168.0.2'
    """
    return dec2addr(addr2dec(ip) + add_value)



class ExcelDataGenerator:

    def __init__(self):
        self._data = {}  # type: Dict[str, Dict[str, str]]

    def load_generator(self, filename: str, sheetname: str = 'Sheet1', key=None):
        """
        加载excel文件, 这个是个迭代器
        @param filename: 文件名
        @param sheetname: sheet名
        @param key: 唯一字段名，如果为None，则按照第一行的字段名称自动生成
        """
        wb = load_workbook(filename)
        sheet = wb[sheetname]

        vars_list = []

        if key is None:
            key = sheet.cell(row=1, column=1).value

        # 取第一行中所有值，生成变量列表
        for col in sheet.iter_cols(min_col=1, max_col=sheet.max_column, min_row=1, max_row=1):
            for cell in col:
                vars_list.append(cell.value)

        # 对每一行进行遍历
        for row in range(3, sheet.max_row + 1):
            idx = 0
            ds = {}
            for col in sheet.iter_cols(min_col=1, max_col=sheet.max_column, min_row=row, max_row=row):
                for cell in col:
                    ds[vars_list[idx]] = cell.value
                    idx += 1
            yield ds

            # 合并到所有
            self._data[ds[key]] = ds

    def load(self, filename: str, sheetname: str = 'Sheet1', key=None):
        """
        加载excel文件, 屏蔽掉了迭代器
        @param filename: 文件名
        @param sheetname: sheet名
        @param key: 唯一字段名，如果为None，则按照第一行的字段名称自动生成
        """
        for ds in self.load_generator(filename, sheetname, key):
            ...

    @property
    def data(self):
        """取从Excel中加载的数据"""
        return self._data.items()


def load_template(template_filename: str) -> str:
    """读取模板文件"""
    with open(template_filename, 'r+', encoding='utf-8', errors='ignore') as f:
        return f.read()


def write_template(template_str: str, output_file: str, data: dict):
    """
    将模板写入文件
    @param template_str: 模板字符串
    @param output_file: 输出文件
    @param data: 数据
    """

    # 判断文件夹是否存在
    dir = os.path.dirname(os.path.join(
        os.path.dirname(sys.argv[0]), output_file))
    if not os.path.exists(dir):
        os.makedirs(dir)

    with open(output_file, 'w+', encoding='utf-8') as f1:
        try:
            f1.write(template_str.format(**data))
        except KeyError as e:
            print(f'模板中变量不存在：{e}')
            exit(1)
